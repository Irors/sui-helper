import aiohttp
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
from .modul import from_sui_amount

class Excel:

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Address'
    sheet['B1'] = 'Balance'

    # Настройка размера столбца A
    column = sheet.column_dimensions[openpyxl.utils.get_column_letter(1)]
    column.width = 16  # Задаем ширину столбца (в пикселях)

    # Настройка размера столбца B
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(2)]
    column_B.width = 16


    """ Декорации Excle """
    # Создаем объект для задания стиля выравнивания и отступов
    alignment_style = Alignment(horizontal='center', vertical='center', indent=3, wrap_text=True)

    # Создаем объект для задания стиля границ
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


    # Делаем все ячейки с жирным шрифтом
    for i in ('AB'):
        cell = sheet[f'{i}1']
        bold_font = Font(bold=True)
        cell.font = bold_font

        # Применяем стиль выравнивания и отступов к ячейке
        cell.alignment = alignment_style
        # Применяем стиль границ
        cell.border = border_style

    workbook.save('result\\Balance.xlsx')


async def fetch(session, number, json_data, address):
    async with session.post('https://internal.suivision.xyz/mainnet/api', json=json_data) as response:
        res = await response.json()
        Excel.sheet[f'A{number}'] = address
        Excel.sheet[f'B{number}'] = from_sui_amount(int(res['result']['totalBalance']))

        Excel.workbook.save('result\\Balance.xlsx')

async def get_balance(client, token, amount, number, address):
    async with aiohttp.ClientSession() as session:
        json_data = {
            'jsonrpc': '2.0',
            'id': 1,
            'method': 'suix_getBalance',
            'params': [
                f'{address}',
                f'{token}',
            ],
        }

        await fetch(session, number, json_data, address)